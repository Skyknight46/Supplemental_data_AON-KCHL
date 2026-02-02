"""
ExcelCsvHandler: Unified handler for Excel (.xlsx) and CSV (.csv) file operations.
Provides methods for reading, writing, and printing file contents.
"""
__author__ = "Tom Ummenthun"
__date__ = "07-10-2025"

import logging
from pathlib import Path
import openpyxl as op
import pandas as pd


class ExcelCsvHandler:
    """
    Provides methods for reading, writing, and printing Excel and CSV files.
    """

    @staticmethod
    def load_excel_workbook(filepath: str | Path) -> op.Workbook:
        """Load an Excel workbook from the specified file path.

        Args:
            filepath (str | Path): Path to the Excel file.

        Returns:
            openpyxl.Workbook: Loaded workbook object.

        Raises:
            FileNotFoundError: If the file does not exist.
            openpyxl.utils.exceptions.InvalidFileException: If the file is not a valid Excel file.
        """
        try:
            return op.load_workbook(filepath)
        except Exception as e:
            logging.error("Failed to load Excel workbook: %s (%s)", filepath, e)
            raise

    @staticmethod
    def get_header_row(
        wb: op.Workbook,
        sheet_name: str,
        header_row: int = 1
    ) -> tuple[str, ...] | None:
        """Return the header row as a tuple from the given sheet, all lowercase."""
        try:
            sheet = wb[sheet_name if sheet_name else wb.active.title]
        except KeyError:
            logging.error("Sheet '%s' not found in workbook.", sheet_name)
            return None
        for row in sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
            return tuple(cell.lower() if isinstance(cell, str) else cell for cell in row)

    @staticmethod
    def update_excel_cell(
        wb: op.Workbook,
        sheet_name: str,
        row: int,
        col: int,
        value: str = ''
    ) -> None:
        """Update a specific cell in the Excel sheet if it is empty.
        If the cell is not empty, return without updating.
        """
        try:
            sheet = wb[sheet_name]
        except KeyError:
            logging.error("Sheet '%s' not found in workbook.", sheet_name)
            return
        cell = sheet.cell(row=row, column=col)
        if cell.value not in (None, ''):
            return
        cell.value = value
        cell.fill = op.styles.PatternFill(
            start_color="ADD8E6",
            end_color="ADD8E6",
            fill_type="solid"
        )

    @staticmethod
    def print_excel_contents(wb: op.Workbook, sheet_name: str) -> None:
        """Print contents of a specific sheet in an Excel file."""
        try:
            ws = wb[sheet_name]
        except KeyError:
            logging.error("Sheet '%s' not found in workbook.", sheet_name)
            return
        header = ExcelCsvHandler.get_header_row(wb, sheet_name, 1)[1:]
        logging.info("Header Row: %s", header)
        for row in ws.iter_rows(values_only=True):
            if row[0] in (None, ''):
                return
            logging.info("%s", row)

    @staticmethod
    def read_csv_into_df(
        filepath: str | Path,
        encoding: str,
        sep: str,
        header_row: int
    ) -> pd.DataFrame:
        """Read a CSV file into a DataFrame."""
        try:
            return pd.read_csv(
                filepath,
                dtype=str,
                encoding=encoding,
                engine='python',
                sep=sep,
                header=header_row,
                skip_blank_lines=False,
                skipfooter=1
            )
        except Exception as e:
            logging.error("Failed to read CSV file: %s (%s)", filepath, e)
            raise

    @staticmethod
    def print_csv_contents(df: pd.DataFrame) -> None:
        """Print contents of a DataFrame."""
        logging.info("Header Row: %s", list(df.columns))
        for _, row in df.iterrows():
            logging.info("%s", tuple(row))

    @staticmethod
    def get_csv_column(col_name: str, df: pd.DataFrame) -> pd.Series:
        """Extract a specific column from a DataFrame."""
        if col_name in df.columns:
            return df[col_name]
        raise ValueError(f"Column '{col_name}' not found in DataFrame.")

    @staticmethod
    def get_csv_row(index: int, df: pd.DataFrame) -> tuple[str, ...]:
        """Extract a specific row from a DataFrame as a tuple of cell values."""
        if 0 <= index < len(df):
            return tuple(df.iloc[index])
        raise ValueError(f"Row index '{index}' out of bounds.")

    @staticmethod
    def load_excel_into_df(
        filepath: str | Path,
        sheetname: str | int = None,
        header_row: int = 0
    ) -> pd.DataFrame:
        """Read an Excel file into a DataFrame."""
        return pd.read_excel(
            filepath,
            sheet_name=sheetname,
            dtype=str,
            header=header_row,
            engine='openpyxl'
        )

    @staticmethod
    def save_excel_workbook(wb: op.Workbook, filepath: str | Path) -> None:
        """Save an Excel workbook to the specified file path."""
        wb.save(filepath)
