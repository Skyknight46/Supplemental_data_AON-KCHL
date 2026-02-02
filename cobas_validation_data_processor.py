"""
GLIMS Cobas Validation Data Processor.

Automates the transfer of test results from GLIMS laboratory information
system CSV exports into Cobas analyzer validation Excel templates. Supports
both non-reproducibility (patient sample) and reproducibility (QC material)
validation workflows.

IMPORTANT: This script requires 'config_data.json' to be present in the
same directory. The config file contains all test mnemonics, QC materials,
and validation parameters.

Glossary of Abbreviations:
    GLIMS: Laboratory Information Management System
    QC: Quality Control
    SKML: Stichting Kwaliteitsbewaking Medische Laboratoriumdiagnostiek
          (Dutch Foundation for Quality Assessment in Medical Laboratories)
    C-module: Chemistry module on Cobas analyzer
    E-module: Electrochemiluminescence (immunoassay) module on Cobas analyzer
    Pro: Cobas Pro integrated analyzer system
    LHB: Lithium Heparin Blood (sample type)
    URP: Urine Pregnancy (sample type)
    EDTA: Ethylenediaminetetraacetic acid (anticoagulant)
    SER: Serum (sample type)
    CEF: Cerebrospinal Fluid
    NAF: Sodium Fluoride (anticoagulant)
    HIV: Human Immunodeficiency Virus
    VLK: Validation control materials prefix

Expected CSV format for non-reproducibility testing:
    - First 4 rows: metadata (skipped)
    - 5th row: empty row (skipped)
    - 6th row: header row with column names
    - 7th row onwards: data rows
    
    Columns:
    - 1st: Patient ID
    - 2nd: Testrun identifier (e.g., 'v_testrun1', 'v_testrun2', 'v_testrun3')
    - 3rd: Either 'Order id' or start of test mnemonics if 'Order id' is absent
    - 4th+: Test result columns (one per test mnemonic)

Expected CSV format for reproducibility testing:
    - First 4 rows: metadata (skipped)
    - 5th row: empty row (skipped)
    - 6th row: header row with column names
    - 7th row onwards: data rows
    
    Columns:
    - 1st: Measurement date (dd-mm-yyyy format)
    - 2nd: QC material identifier
    - 3rd: Analyzer identifier
    - 4th+: Test result columns (one per test mnemonic)
"""
__author__ = "Tom Ummenthun"
__date__ = "22-10-2025"

import json
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from typing import Dict, List, Optional, Set, Tuple

from dateutil import parser
import openpyxl
from pandas import DataFrame

from excel_csv_handler import ExcelCsvHandler

# Constants for CSV parsing
CSV_METADATA_ROWS = 4  # Number of metadata rows to skip
CSV_EMPTY_ROW = 5  # Empty row position
CSV_HEADER_ROW_NUMBER = 6  # Header row position (1-indexed for documentation)
CSV_SKIPROWS = 5  # Number of rows to skip before header (0-indexed for pandas)
CSV_DATA_START_ROW = 7  # First data row position

# Constants for Excel worksheet indices
C_MODULE_SHEET_INDEX = 0  # Chemistry module worksheet
E_MODULE_SHEET_INDEX = 1  # Electrochemiluminescence module worksheet

# Constants for column offsets based on analyzer type
C_MODULE_BASE_OFFSET_8000 = 8  # Base column for Cobas 8000
C_MODULE_BASE_OFFSET_PRO = 9  # Base column for Cobas Pro
E_MODULE_BASE_OFFSET_8000 = 14  # Base column for Cobas 8000
E_MODULE_BASE_OFFSET_PRO = 15  # Base column for Cobas Pro
DEFAULT_TESTRUN_OFFSET = 40  # Default offset for unknown testruns

# Constants for reproducibility processing
REPRO_FIRST_DATA_ROW = 3  # First row for reproducibility data
REPRO_HEADER_ROW = 2  # Header row in reproducibility sheets
EXCEL_NAME_COLUMN = 1  # Column index for test names (0-indexed)
EXCEL_ROW_START = 2  # Starting row for data lookup

# File encoding for Windows regional settings
DEFAULT_ENCODING = 'windows-1252'

# Load configuration data from JSON file
def load_config() -> Dict:
    """
    Load configuration data from config_data.json file.
    
    The config file must be located in the same directory as this script.
    Uses __file__ to determine the script's location and constructs the path
    to config_data.json relative to the script directory.
    
    Returns:
        Dict: Configuration dictionary containing:
            - offset_map: Column offsets for different testruns
            - pro_stations: List of Cobas Pro station identifiers
            - test_mnemonics: Dict with c_module and e_module test codes
            - qc_material: QC material mappings (deprecated/old names)
            - qc_material_pro: QC material mappings (current/unified names)
            - skip_hiv: List of HIV QC materials to skip during processing
    
    Raises:
        SystemExit: Exits the program with an error message if config 
            cannot be loaded or is invalid.
    """
    try:
        config_path = os.path.join(os.path.dirname(__file__), 
                                   'config_data.json')
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        # Validate required keys
        required_keys = ['offset_map', 'pro_stations', 'test_mnemonics',
                        'qc_material', 'qc_material_pro', 'skip_hiv']
        missing_keys = [key for key in required_keys if key not in config]
        if missing_keys:
            raise KeyError(
                f"Missing required keys in config: {', '.join(missing_keys)}"
            )
        
        return config
    except FileNotFoundError:
        messagebox.showerror(
            "Configuration Error",
            f"Cannot find 'config_data.json' in the script directory.\n\n"
            f"Expected location:\n{os.path.dirname(__file__)}\n\n"
            f"Please ensure config_data.json is in the same folder as this "
            f"script."
        )
        sys.exit(1)
    except json.JSONDecodeError as e:
        messagebox.showerror(
            "Configuration Error",
            f"Error parsing config_data.json:\n{str(e)}\n\n"
            f"Please check that the JSON file is valid."
        )
        sys.exit(1)
    except KeyError as e:
        messagebox.showerror(
            "Configuration Error",
            f"Invalid config_data.json:\n{str(e)}"
        )
        sys.exit(1)
    except Exception as e:
        messagebox.showerror(
            "Configuration Error",
            f"Unexpected error loading config_data.json:\n{str(e)}"
        )
        sys.exit(1)

# Load all configuration data
CONFIG = load_config()
OFFSET_MAP = CONFIG['offset_map']
PRO_STATIONS = CONFIG['pro_stations']
TEST_MNEMONICS = CONFIG['test_mnemonics']
QC_MATERIAL = CONFIG['qc_material']
QC_MATERIAL_PRO = CONFIG['qc_material_pro']
SKIP_HIV = set(CONFIG['skip_hiv'])

def choose_file(title: str, filetype: str = None) -> str:
    """
    Open a file dialog to choose a file of the specified type.

    Args:
        title (str): Dialog title for the file dialog window.
        filetype (str, optional): File extension to filter by (e.g., '.csv').

    Returns:
        str: The path to the selected file, or an empty string if cancelled.
    """
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    if filetype:
        filetype = filetype if filetype.startswith('.') else f'.{filetype}'
        file_path = filedialog.askopenfilename(
            title=f"Select the {title}{filetype} file",
            filetypes=[(f"{filetype[1:]} files", filetype)]
        )
    else:
        file_path = filedialog.askopenfilename(title="Select a file")
    return file_path


def get_testrun_column_offset(testrun: str, csv_row_index: int) -> int:
    """
    Get the column offset for a given testrun identifier.
    
    Maps testrun identifiers (v_testrun1, v_testrun2, v_testrun3) to their
    respective column offsets for writing data into the Excel template.

    Args:
        testrun: The testrun identifier (e.g., 'v_testrun1').
        csv_row_index: Row index in the CSV for error reporting.

    Returns:
        The column offset for the testrun, or DEFAULT_TESTRUN_OFFSET if unknown.
    """
    if testrun in OFFSET_MAP:
        return OFFSET_MAP[testrun]
    
    print(
        f"Warning: Unknown testrun ID '{testrun}' at CSV row "
        f"{csv_row_index} - using default offset {DEFAULT_TESTRUN_OFFSET}."
    )
    return DEFAULT_TESTRUN_OFFSET


def get_sheet_and_column_base(
    mnemonic: str,
    patient_id: str,
    test_codes_map: Dict[str, List[str]]
) -> Tuple[Optional[int], Optional[int]]:
    """
    Determine the worksheet index and base column for a given test mnemonic.
    
    Identifies whether the test belongs to the C-module (chemistry) or
    E-module (immunoassay) and returns the appropriate worksheet index
    and base column offset. The base column varies depending on whether
    the analyzer is a Cobas 8000 or Cobas Pro.

    Args:
        mnemonic: Test mnemonic code to look up.
        patient_id: Patient identifier (used to distinguish analyzer type:
            '8000' in ID indicates Cobas 8000, otherwise Cobas Pro).
        test_codes_map: Dictionary mapping 'c_module' and 'e_module' to
            their respective test code lists.

    Returns:
        Tuple of (worksheet_index, base_column_index), or (None, None)
        if the mnemonic is not found in either module.
    """
    mnemonic_lower = mnemonic.lower()
    c_module_codes = [code.lower() for code in test_codes_map['c_module']]
    e_module_codes = [code.lower() for code in test_codes_map['e_module']]
    
    if mnemonic_lower in c_module_codes:
        base_column = (
            C_MODULE_BASE_OFFSET_8000 if '8000' in patient_id
            else C_MODULE_BASE_OFFSET_PRO
        )
        return (C_MODULE_SHEET_INDEX, base_column)
    
    if mnemonic_lower in e_module_codes:
        base_column = (
            E_MODULE_BASE_OFFSET_8000 if '8000' in patient_id
            else E_MODULE_BASE_OFFSET_PRO
        )
        return (E_MODULE_SHEET_INDEX, base_column)
    
    print(f"Warning: Unknown mnemonic '{mnemonic}' - skipping entry.")
    return (None, None)


def find_test_row_in_worksheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    mnemonic: str
) -> Optional[int]:
    """
    Find the row index for a test mnemonic in the Excel worksheet.
    
    Searches the worksheet starting from row 2 to find a row where the
    test name (in column B, index 1) matches the given mnemonic.

    Args:
        worksheet: Excel worksheet object to search.
        mnemonic: Test mnemonic code to find in the worksheet.

    Returns:
        Row index (1-based Excel numbering) if found, None otherwise.
    """
    for row_index, worksheet_row in enumerate(
        worksheet.iter_rows(
            min_row=EXCEL_ROW_START,
            max_row=worksheet.max_row
        ),
        start=EXCEL_ROW_START
    ):
        # Test name is in column B (index 1)
        test_name_cell = worksheet_row[EXCEL_NAME_COLUMN]
        if (test_name_cell.value and
            str(test_name_cell.value).strip().lower() == mnemonic.lower()):
            return row_index
    
    return None

def process_skml(glims_df: DataFrame, workbook: openpyxl.Workbook) -> None:
    """
    Process non-reproducibility validation data from GLIMS output.
    
    Reads test results from the GLIMS CSV data and writes them into the
    appropriate cells in the Excel validation template. Handles both
    C-module (chemistry) and E-module (immunoassay) test results.

    Args:
        glims_df: DataFrame containing GLIMS output data.
        workbook: Excel workbook object to update with test results.
    """
    # Debug: Print column names to diagnose issue
    print(f"Column names: {list(glims_df.columns)}")
    print(f"Column 2 (index 2): '{glims_df.columns[2]}'")
    print(f"Column 2 starts with 'k_': {glims_df.columns[2].startswith('k_')}")
    
    # Determine if 'Order id' column is present
    # If column 2 starts with 'k_', it's a test mnemonic (no order_id)
    # If column 2 doesn't start with 'k_', it's the order_id column
    has_order_column = not str(glims_df.columns[2]).lower().startswith("k_")
    
    print(f"Has order column: {has_order_column}")
    print(f"Mnemonic start column: {3 if has_order_column else 2}")

    # Build list of test mnemonics from column headers
    mnemonic_start_col = 3 if has_order_column else 2
    test_mnemonics = [
        col.lower() for col in glims_df.columns[mnemonic_start_col:]
    ]
    
    debug_mode = messagebox.askyesno(
        "Debug Mode", "Enable debug prints?"
    )
    
    for row in glims_df.itertuples(index=False):
        patient_id = row[0]
        testrun_id = str(row[1]).lower()
        
        if testrun_id not in OFFSET_MAP:
            print(
                f"Warning: Unknown testrun ID '{testrun_id}' - skipping row."
            )
            continue
        
        for test_mnemonic in test_mnemonics:
            mnemonic_column_index = (
                test_mnemonics.index(test_mnemonic) + mnemonic_start_col
            )
            test_result_value = row[mnemonic_column_index]
            
            if test_result_value == '-' or test_result_value is None:
                continue
            
            worksheet_index, base_column_index = get_sheet_and_column_base(
                test_mnemonic,
                patient_id,
                TEST_MNEMONICS
            )
            
            if worksheet_index is None or base_column_index is None:
                continue
            
            worksheet = workbook.worksheets[worksheet_index]
            excel_row_index = find_test_row_in_worksheet(
                worksheet, test_mnemonic
            )
            result_column_index = (
                base_column_index + OFFSET_MAP[testrun_id]
            )
            
            # Convert decimal separator for regional settings
            value_to_write = (
                str(test_result_value).replace('.', ',')
                if isinstance(test_result_value, str)
                else test_result_value
            )
            
            if excel_row_index is None or result_column_index is None:
                print(
                    f"Warning: Row/column not found for mnemonic "
                    f"'{test_mnemonic}' (Patient ID: '{patient_id}') - "
                    f"skipping entry."
                )
                continue
            
            ExcelCsvHandler.update_excel_cell(
                workbook,
                worksheet.title,
                excel_row_index,
                result_column_index,
                value_to_write
            )
            
            if debug_mode:
                with open("debug_skml_log.txt", "a",
                         encoding=DEFAULT_ENCODING) as debug_file:
                    debug_file.write(
                        f"Processing mnemonic '{test_mnemonic}' "
                        f"for patient ID '{patient_id}':\n"
                        f"  GLIMS column index: {mnemonic_column_index}\n"
                        f"  Sheet name: {worksheet.title}\n"
                        f"  Row index: {excel_row_index}\n"
                        f"  Value column index: {result_column_index}\n"
                        f"  Value to write: {test_result_value}\n"
                    )


def determine_repro_sheet(
    pro_validation: bool,
    qc_material: str,
    analyser: str
) -> str:
    """
    Determine the reproducibility worksheet for a QC material.
    
    Maps QC material identifiers to the appropriate reproducibility
    worksheet ('REPRO controle 1', '2', or '3') based on material type
    and analyzer configuration.

    Args:
        pro_validation: Whether this is Cobas Pro validation.
        qc_material: QC material identifier.
        analyser: Analyzer identifier.

    Returns:
        Name of the reproducibility worksheet.

    Raises:
        ValueError: If the QC material or analyser is not recognized.
    """
    # Special handling for interference materials
    if qc_material in {"VLK_Hemolyse", "VLK_Ict-V", "VLK_Lip-V"}:
        if analyser in {"VLK_PRO-4_C503", "VLK_PRO-3_C503"}:
            return "REPRO controle 1"
        if analyser in {"VLK_PRO-4_C703", "VLK_PRO-3_C703"}:
            return "REPRO controle 2"
        if analyser == "VLK_C500PRO-01":
            # Deprecated non-Pro logic
            return "REPRO controle 3"
        raise ValueError(
            f"Unknown analyser '{analyser}' for special QC material: "
            f"{qc_material}"
        )
    
    if pro_validation:
        if (qc_material in QC_MATERIAL_PRO['low'] or
            qc_material in QC_MATERIAL['low']):
            return "REPRO controle 1"
        if (qc_material in QC_MATERIAL_PRO['high'] or
            qc_material in QC_MATERIAL['high']):
            return "REPRO controle 2"
        if any(marker in qc_material for marker in ["VLK_Bil", "VLK_Vrij"]):
            return "REPRO controle 3"
        raise ValueError(f"Unknown QC material: {qc_material}")
    
    # Deprecated non-Pro validation path
    raise ValueError("Non-Pro validation is deprecated")

def get_repro_test_column(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    mnemonic: str
) -> Optional[int]:
    """
    Find the column index for a test mnemonic in reproducibility sheet.
    
    Searches the header row (row 2) of the reproducibility worksheet to
    find the column containing the specified test mnemonic.

    Args:
        worksheet: Excel worksheet object to search.
        mnemonic: Test mnemonic code to find in the header row.

    Returns:
        Column index (1-based) if found, None otherwise.
    """
    for column_index in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=REPRO_HEADER_ROW, 
                                   column=column_index).value
        if (cell_value is not None and
            cell_value.lower() == mnemonic.lower()):
            return column_index
    
    return None

def update_repro_row(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    row_index: int,
    column_index: int,
    measurement_data: List
) -> None:
    """
    Write a measurement value to the reproducibility worksheet.
    
    Updates a specific cell in the reproducibility worksheet with a test
    result value, converting decimal separators for regional settings.

    Args:
        workbook: Excel workbook object.
        sheet_name: Name of the worksheet to update.
        row_index: Row index to update (1-based Excel numbering).
        column_index: Column index to update (1-based Excel numbering).
        measurement_data: List containing [datetime, value, qc_material, 
            analyser].
    """
    if sheet_name not in workbook.sheetnames:
        print(
            f"Warning: Sheet '{sheet_name}' does not exist in workbook - "
            f"skipping update for row {row_index}, column {column_index}."
        )
        return
    
    if column_index is None:
        print(
            f"Warning: Column is None for update in sheet '{sheet_name}', "
            f"row {row_index} - skipping update."
        )
        return
    
    # Convert decimal separator for regional settings
    value_to_write = (
        str(measurement_data[1]).replace('.', ',')
        if isinstance(measurement_data[1], str)
        else measurement_data[1]
    )
    
    ExcelCsvHandler.update_excel_cell(
        workbook,
        sheet_name,
        row_index,
        column_index,
        value_to_write
    )

def process_reproducibility(
    glims_df: DataFrame,
    workbook: openpyxl.Workbook
) -> None:
    """
    Process reproducibility validation data from GLIMS output.
    
    Reads QC material test results, sorts them chronologically, and writes
    them to the appropriate reproducibility worksheets in the Excel template.

    Args:
        glims_df: DataFrame containing GLIMS reproducibility data.
        workbook: Excel workbook object to update with QC results.
    """
    # Cobas Pro validation is currently the standard
    is_pro_validation = True
    
    test_mnemonics = [
        column.lower() for column in glims_df.columns.tolist()[3:]
    ]
    
    measurements_by_test = build_repro_measurements(
        glims_df, test_mnemonics, is_pro_validation
    )
    write_repro_results(measurements_by_test, workbook, is_pro_validation)

def build_repro_measurements(
    glims_df: DataFrame,
    test_mnemonics: List[str],
    is_pro_validation: bool
) -> Dict[str, List]:
    """
    Build measurement data structure from reproducibility CSV.
    
    Organizes QC measurements by test mnemonic, filtering by analyzer
    number for Pro validation and excluding HIV materials that should
    be skipped.

    Args:
        glims_df: DataFrame containing GLIMS reproducibility data.
        test_mnemonics: List of test mnemonic codes from CSV headers.
        is_pro_validation: Whether this is Cobas Pro validation.

    Returns:
        Dictionary mapping test mnemonics to lists of measurement data.
        Each measurement is [datetime, value, qc_material, analyser].
    """
    analyser_number = None
    
    if is_pro_validation:
        analyser_number = prompt_for_analyser_number()
    
    measurements = {mnemonic: [] for mnemonic in test_mnemonics}
    
    for row in glims_df.itertuples(index=False):
        try:
            measurement_datetime = parser.parse(str(row[0]), dayfirst=True)
        except parser.ParserError:
            if row[0] is not None:
                print(
                    f"Warning: Invalid date format '{row[0]}' - skipping row."
                )
            continue
        
        qc_material = row[1]
        analyser = row[2]
        
        # Skip HIV materials that are not used
        if qc_material in SKIP_HIV:
            continue
        
        # Filter by analyzer number for Pro validation
        if is_pro_validation and analyser_number is not None:
            if f"PRO-{analyser_number}" not in analyser:
                continue
        
        # Process all test values in this row
        for index, value in enumerate(row[3:]):
            if value != '-' and value is not None:
                measurements[test_mnemonics[index]].append(
                    [measurement_datetime, value, qc_material, analyser]
                )
    
    return measurements


def prompt_for_analyser_number() -> Optional[int]:
    """
    Prompt user to select Cobas Pro analyzer number (3 or 4).
    
    Shows a dialog asking the user to specify which Cobas Pro analyzer
    (Pro 3 or Pro 4) is being validated. Allows up to 3 retry attempts.

    Returns:
        Analyzer number (3 or 4), or None if user cancels or exceeds retries.
    """
    retry_limit = 3
    
    for retry in range(retry_limit):
        try:
            user_input = simpledialog.askstring(
                "Analyser Name",
                "Cobas pro 3 or pro 4? (Enter just the number)"
            )
            
            if user_input is None:
                return None
            
            analyser_number = int(user_input)
            
            if analyser_number in {3, 4}:
                return analyser_number
            
            print(
                f"Warning: Invalid analyzer number {analyser_number}. "
                f"Please enter 3 or 4."
            )
        except (TypeError, ValueError):
            print(
                f"Warning: Invalid input. Please enter '3' or '4'. "
                f"Attempt {retry + 1}/{retry_limit}"
            )
    
    print(
        f"Error: Failed to get valid analyzer number after {retry_limit} "
        f"attempts."
    )
    return None

def write_repro_results(
    measurements_by_test: Dict[str, List],
    workbook: openpyxl.Workbook,
    is_pro_validation: bool
) -> None:
    """
    Write sorted reproducibility measurements to Excel worksheets.
    
    Sorts measurements chronologically and writes them sequentially to
    the appropriate reproducibility control worksheets, handling special
    cases for HIV tests.

    Args:
        measurements_by_test: Dictionary mapping test mnemonics to lists
            of measurement data.
        workbook: Excel workbook object to update.
        is_pro_validation: Whether this is Cobas Pro validation.
    """
    for test_name, measurements in measurements_by_test.items():
        # Sort measurements chronologically
        measurements.sort(key=lambda x: x[0])
        
        # Track row indices for each control worksheet
        row_control_1 = REPRO_FIRST_DATA_ROW
        row_control_2 = REPRO_FIRST_DATA_ROW
        
        for measurement in measurements:
            datetime_val, value, qc_material, analyser = measurement
            
            sheet_name = determine_repro_sheet(
                is_pro_validation,
                qc_material,
                analyser
            )
            
            # Special logic for HIV tests with multiple QC materials
            sheet_name = apply_hiv_sheet_override(
                test_name, qc_material, sheet_name
            )
            
            column_index = get_repro_test_column(workbook[sheet_name],
                                                 test_name)
            
            if column_index is None:
                print(
                    f"Warning: Test mnemonic '{test_name}' not found in "
                    f"sheet '{sheet_name}' - skipping entry."
                )
                continue
            
            if sheet_name == "REPRO controle 1":
                update_repro_row(
                    workbook,
                    sheet_name,
                    row_control_1,
                    column_index,
                    measurement
                )
                row_control_1 += 1
            elif sheet_name == "REPRO controle 2":
                update_repro_row(
                    workbook,
                    sheet_name,
                    row_control_2,
                    column_index,
                    measurement
                )
                row_control_2 += 1


def apply_hiv_sheet_override(
    test_name: str,
    qc_material: str,
    default_sheet: str
) -> str:
    """
    Apply special HIV test routing logic.
    
    HIV tests have specific QC material to worksheet mappings that
    override the standard low/high material assignment.

    Args:
        test_name: Test mnemonic code (lowercase).
        qc_material: QC material identifier.
        default_sheet: Default sheet name from standard logic.

    Returns:
        Sheet name to use (may be overridden for HIV tests).
    """
    hiv_mappings = {
        ("m_ahiv_eclia_ser", "VLK_PC_HIV_3"): "REPRO controle 1",
        ("m_hivag_eclia_ser", "VLK_PC_HIV_3"): "REPRO controle 2",
        ("m_ahiv_eclia_ser", "VLK_PC_HIV_5"): "REPRO controle 2",
        ("m_hivag_eclia_ser", "VLK_PC_HIV_5"): "REPRO controle 1",
    }
    
    return hiv_mappings.get((test_name, qc_material), default_sheet)

def main() -> None:
    """
    Main entry point for the GLIMS Cobas validation processor.
    
    Prompts user for file selections and processing options, then loads
    data and routes to appropriate validation workflow (reproducibility
    or non-reproducibility).
    """
    # Determine validation workflow
    is_reproducibility = messagebox.askyesno(
        "Reproducibility",
        "Is this for reproducibility testing?"
    )
    
    # Get file paths from user
    glims_csv_path = choose_file("Select the GLIMS output", ".csv")
    if not glims_csv_path:
        sys.exit(0)
    
    template_excel_path = choose_file(
        "Select the Validation Excel file", ".xlsx"
    )
    if not template_excel_path:
        sys.exit(0)
    
    # Get CSV separator
    csv_separator = prompt_for_csv_separator()
    
    # Load data files
    glims_dataframe = load_glims_csv(
        glims_csv_path,
        csv_separator,
        CSV_SKIPROWS
    )
    
    validation_workbook = ExcelCsvHandler.load_excel_workbook(
        template_excel_path
    )
    
    # Process data based on workflow type
    if is_reproducibility:
        process_reproducibility(glims_dataframe, validation_workbook)
    else:
        process_skml(glims_dataframe, validation_workbook)
    
    # Save results
    save_path = filedialog.asksaveasfilename(
        title="Save the updated Excel file as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    if save_path:
        ExcelCsvHandler.save_excel_workbook(validation_workbook, save_path)
        print(f"Successfully saved validation results to: {save_path}")
    else:
        print("Warning: File save cancelled. Results not saved.")


def prompt_for_csv_separator() -> str:
    """
    Prompt user for CSV file separator character.
    
    Returns:
        CSV separator character (defaults to ';' if user cancels or provides
        empty input).
    """
    separator_input = simpledialog.askstring(
        "Separator",
        "Enter the CSV separator (defaults to ';'):\t\t",
        initialvalue=';'
    )
    
    return separator_input if separator_input else ';'


def load_glims_csv(
    file_path: str,
    separator: str,
    skiprows: int
) -> DataFrame:
    """
    Load GLIMS CSV file into a DataFrame.
    
    Loads the CSV with Windows-1252 encoding to handle special characters,
    skips metadata rows, and converts column headers to lowercase.

    Args:
        file_path: Path to the GLIMS CSV file.
        separator: CSV separator character.
        skiprows: Number of rows to skip before the header row (0-indexed).

    Returns:
        DataFrame with GLIMS data and lowercase column names.
    """
    dataframe = ExcelCsvHandler.read_csv_into_df(
        file_path,
        DEFAULT_ENCODING,
        separator,
        skiprows
    )
    
    # Convert headers to lowercase for consistent processing
    dataframe.columns = [col.lower() for col in dataframe.columns]
    
    return dataframe


if __name__ == "__main__":
    main()
